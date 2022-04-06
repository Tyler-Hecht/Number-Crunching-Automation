import os
import sys
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors

'''
Checks if the program is running as an executable or if the program is running through an IDE
Will return the application path of the source file
'''
def get_application_path():
    application_path = ' '
    if getattr(sys, 'frozen', False):#if the program is running as an executable
        application_path = os.path.dirname(sys.executable)
    else: #if program is running through ide
        application_path = os.path.dirname(os.path.abspath(__file__))
    return application_path
'''
Calculate the difference between column B and C in the coder files
'''
def calculate_difference(application_path):
    files = os.listdir(application_path + "/Input/") #get all files in Input folder
    for file in files:
        if(file == ".DS_Store"): #skip .DS_Store file
            continue
        wb = openpyxl.load_workbook(application_path + "/Input/" + file) #load excel file
        for sheet in wb: #iterate through each sheet in excel file
            if sheet.title == 'AVERAGES ACROSS CODERS':
                continue
            row_count = 1
            for row in sheet: #for each row in sheet
                difference = 0
                if row[1].value == None and row[2].value == None: #if both row entry are empty
                    difference 
                elif row[1].value == None: #if the first row entry is empty
                    difference = row[2].value - 0
                elif row[2].value == None: #if the second row entry is empty
                    difference = 0 - row[1].value
                else:
                    difference = row[2].value - row[1].value
                sheet.cell(row=row_count, column=4).value = difference #add the difference to the fourth column
                row_count += 1 #continue to the next row
        wb.save(application_path + "/Input/" + file) #save the excel file
'''
 Insert the column header for the coder files
'''       
def insert_col_header(application_path):
    files = os.listdir(application_path + "/Input/")
    for file in files:
        if(file == ".DS_Store"):
            continue
        wb = openpyxl.load_workbook(application_path + "/Input/" + file)
        col_list = ['Left Longest', 'Right Longest', 'Left', 'Right', 'Center', 'Total Look', 'Trial Length', 'Attention']
        col_count = 5
        list_count = 0
        for sheet in wb:
            if sheet.title == 'AVERAGES ACROSS CODERS':
                continue
            while col_count <= 12:#iterate through all the column title in col_list
                sheet.cell(row=2, column=col_count).value = col_list[list_count]#enter value
                col_count += 1
                list_count += 1
            col_count = 5
            list_count = 0
        wb.save(application_path + "/Input/" + file)
'''
Compute longest looks, sum of each looks, total look, total trial, and attention
Will skip the first 30 seconds of each trial
'''
def compute(application_path):
    files = os.listdir(application_path + "/Input/") #get all files of input folder
    for file in files: #iterate through all file
        if(file == ".DS_Store"): #skip .DS_Store file
            continue
        wb = openpyxl.load_workbook(application_path + "/Input/" + file) #load the excel file
        coder_num = 0 #coder 1 or coder 2
        for sheet in wb.worksheets: #for each sheet in excel file
            coder_num += 1
            if sheet.title == 'AVERAGES ACROSS CODERS': #skip average across coder file
                continue
            row_count = 3 #start row
            l_max = 0 #left longest
            r_max = 0 #right longest
            l_sum = 0 #right 
            r_sum = 0 #left
            c_sum = 0 #center
            start_time = None #start time
            end_time = None #end time
            excel_row_count = 1

            #EDITED FOR FACETALK: immediately calculates time at start of trial (with no buffer of 42 frames)
            for row in sheet.iter_rows():
                start_look = row[1].value #start value for look
                end_look = row[2].value #end value for look
                if row[0].value == None: #if no more values then break from loop
                    break
                code = row[0].value.strip().upper() #uppercase all code
                if code == 'B': #if start of trial
                    #count += 1
                    start_time = row[1].value #get start time
                    #print(rt_max, count)
                elif code == 'L': #if left look
                    if start_look == None:
                        print('ERROR: Missing onset time in row', excel_row_count, 'Coder', coder_num, 'in file', file)
                    if end_look == None:
                        print('ERROR: Missing offset time in row', excel_row_count, 'Coder', coder_num, 'in file', file)
                    if row[3].value > l_max: #check if it is the longest look
                        l_max = row[3].value
                    l_sum = l_sum + row[3].value#add to sum
                elif code == 'R': #if right look
                    if start_look == None:
                        print('ERROR: Missing onset time in row', excel_row_count, 'Coder', coder_num, 'in file', file)
                    if end_look == None:
                        print('ERROR: Missing offset time in row', excel_row_count, 'Coder', coder_num, 'in file', file)
                    if row[3].value > r_max: #check if it is the longest look
                        r_max = row[3].value
                    r_sum = r_sum + row[3].value #add to sum
                elif code == 'C': #if center look
                    if start_look == None:
                        print('ERROR: Missing onset time in row', excel_row_count, 'Coder', coder_num, 'in file', file)
                    if end_look == None:
                        print('ERROR: Missing offset time in row', excel_row_count, 'Coder', coder_num, 'in file', file)
                    c_sum = c_sum + row[3].value #add to sum
                elif code == 'S': #if end of trial
                    end_time = row[1].value #get end time
                    total_look = l_sum + r_sum + c_sum #calculate total_look
                    trial_length = end_time - start_time #calculate trial length
                    attention = total_look / trial_length #calculate attention
                    col_count = 5
                    list_count = 0
                    data = [l_max, r_max, l_sum, r_sum, c_sum, total_look, trial_length, attention]
                    while col_count <= 12: #add calculated data
                        sheet.cell(row=row_count, column=col_count).value = data[list_count]
                        col_count += 1
                        list_count += 1
                    row_count += 1 #go to next row
                    l_max = 0 #initalize all values to 0
                    r_max = 0
                    l_sum = 0
                    r_sum = 0
                    c_sum = 0
                    start_time = None
                    end_time = None
                else:
                    print("Error occuring in row", excel_row_count, "Coder", coder_num, 'in', file)
                excel_row_count += 1
        wb.save(application_path + "/Input/" + file)
'''
Insert column headers for AVERAGE ACROSS CODERS excel sheet
'''
def insert_dis_col_header(application_path):
    files = os.listdir(application_path + "/Input/")
    for file in files:
        if(file == ".DS_Store"):
            continue
        wb = openpyxl.load_workbook(application_path + "/Input/" + file)
        col_list = ['Left Longest', 'Left Discrep','Right Longest', 'Rt Discrep', 'Left',
                    'L dis', 'Right', 'R dis', 'Center', 'C dis','Total Look', 'Total discr', 'Trial Length',
                    'Length Discr', 'Attention']
        col_count = 1
        list_count = 0
        if len(wb.worksheets) == 2:#check if there is only two sheets
            wb.create_sheet('AVERAGES ACROSS CODERS')#create new sheet if it doesn't exisit 
        yellowFill = PatternFill(start_color = 'FFFF00',
                                 end_color = 'FFFF00',
                                 fill_type = 'solid')
        for sheet in wb:
            if sheet.title == 'AVERAGES ACROSS CODERS':
                while col_count <= 15:#add column headers to excel file
                    sheet.cell(row=1, column=col_count).value = col_list[list_count]
                    cur_cell = sheet.cell(1, col_count)
                    if col_count % 2 == 0:
                        cur_cell.fill = yellowFill
                    col_count += 1
                    list_count += 1
                col_count = 1
                list_count = 0
        wb.save(application_path + "/Input/" + file)
'''
Calculate discrepencies between the two coders
'''       
def compute_dis(application_path):
    files = os.listdir(application_path + "/Input/")
    for file in files:
        if(file == ".DS_Store"):
            continue
        wb = openpyxl.load_workbook(application_path + "/Input/" + file)
        sheet1 = wb.worksheets[0] #coder 1
        sheet2 = wb.worksheets[1] #coder 2
        sheet3 = wb.worksheets[2] #average across coders
        l_max = 0 #left longest
        l_max_dis = 0 #left longest disp
        r_max = 0 #right longest
        r_max_dis = 0 #right longest disp
        l_sum = 0 #left 
        l_sum_dis = 0 #left disp
        r_sum = 0 #right 
        r_sum_dis = 0 #right disp 
        c_sum = 0 #center
        c_sum_dis = 0 #center disp
        total_look_sum = 0 #total look
        total_look_dis = 0 #total look disp
        trial_length_sum = 0 #trial length 
        trial_length_dis = 0 #trial length disp
        attention_sum = 0 #attenion
        prev_row_count = 1
        new_row_count = 2
        yellowFill = PatternFill(start_color = 'FFFF00',
                                 end_color = 'FFFF00',
                                 fill_type = 'solid')
        redFill = PatternFill(start_color = 'FF0000',
                              end_color = 'FF0000',
                              fill_type = 'solid')
        for row1, row2 in zip(sheet1, sheet2):
            if prev_row_count == 1 or prev_row_count == 2:
                prev_row_count += 1
                continue
            if row1[4].value == None or row2[4].value == None:
                break
            #calculate left longest and left longest disp
            l_max = (row1[4].value + row2[4].value) / 2
            sheet3.cell(row=new_row_count, column = 1).value = l_max
            l_max_dis = row1[4].value - row2[4].value
            sheet3.cell(row=new_row_count, column = 2).value = l_max_dis
            cur_cell = sheet3.cell(new_row_count, 2)
            if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                cur_cell.fill = redFill
            else:   
                cur_cell.fill = yellowFill
            #calculate right longest and right longest disp
            r_max = (row1[5].value + row2[5].value) / 2
            sheet3.cell(row=new_row_count, column = 3).value = r_max
            r_max_dis = row1[5].value - row2[5].value
            sheet3.cell(row=new_row_count, column = 4).value = r_max_dis
            cur_cell = sheet3.cell(new_row_count, 4)
            if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                cur_cell.fill = redFill
            else:   
                cur_cell.fill = yellowFill
            #calculate left and left disp
            l_sum = (row1[6].value + row2[6].value) / 2
            sheet3.cell(row=new_row_count, column = 5).value = l_sum
            l_sum_dis = row1[6].value - row2[6].value
            sheet3.cell(row=new_row_count, column = 6).value = l_sum_dis
            cur_cell = sheet3.cell(new_row_count, 6)
            if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                cur_cell.fill = redFill
            else:   
                cur_cell.fill = yellowFill
            #calculate right and right disp
            r_sum = (row1[7].value + row2[7].value) / 2
            sheet3.cell(row=new_row_count, column = 7).value = r_sum
            r_sum_dis = row1[7].value - row2[7].value
            sheet3.cell(row=new_row_count, column = 8).value = r_sum_dis
            cur_cell = sheet3.cell(new_row_count, 8)
            if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                cur_cell.fill = redFill
            else:   
                cur_cell.fill = yellowFill
            #calculate center and center disp
            c_sum = (row1[8].value + row2[8].value) / 2
            sheet3.cell(row=new_row_count, column = 9).value = c_sum
            c_sum_dis = row1[8].value - row2[8].value
            sheet3.cell(row=new_row_count, column = 10).value = c_sum_dis
            cur_cell = sheet3.cell(new_row_count, 10)
            if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                cur_cell.fill = redFill
            else:   
                cur_cell.fill = yellowFill
            #calculate total look and total look disp
            total_look_sum = (row1[9].value + row2[9].value) / 2
            sheet3.cell(row=new_row_count, column = 11).value = total_look_sum
            total_look_dis = row1[9].value - row2[9].value
            sheet3.cell(row=new_row_count, column = 12).value = total_look_dis
            cur_cell = sheet3.cell(new_row_count, 12)
            if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                cur_cell.fill = redFill
            else:   
                cur_cell.fill = yellowFill
            #calculate trial length and trial length disp
            trial_length_sum = (row1[10].value + row2[10].value) / 2
            sheet3.cell(row=new_row_count, column = 13).value = trial_length_sum
            trial_length_dis = row1[10].value - row2[10].value
            sheet3.cell(row=new_row_count, column = 14).value = trial_length_dis
            cur_cell = sheet3.cell(new_row_count, 14)
            if cur_cell.value > 15 or cur_cell.value < -15:#color read if dis is greater than 15 or less than -15
                cur_cell.fill = redFill
            else:   
                cur_cell.fill = yellowFill
            
            attention_sum = (row1[11].value + row2[11].value) / 2
            sheet3.cell(row=new_row_count, column = 15).value = attention_sum
            #initalize values to 0
            l_max = 0 #
            l_max_dis = 0 #
            r_max = 0 #
            r_max_dis = 0 #
            l_sum = 0 #
            l_sum_dis = 0 #
            r_sum = 0 #
            r_sum_dis = 0 #
            c_sum = 0 #
            c_sum_dis = 0 #
            total_look_sum = 0 #
            total_look_dis = 0 #
            trial_length_sum = 0
            trial_length_dis = 0
            attention_sum = 0
            new_row_count += 1
        wb.save(application_path + "/Input/" + file)





def main():
    application_path = get_application_path()
    calculate_difference(application_path)
    insert_col_header(application_path)
    compute(application_path)
    insert_dis_col_header(application_path)
    compute_dis(application_path)
    print('Averages across coders completed!')

if __name__ == "__main__":
    main()